# log_watcher.py

import os
import re
import json
import time
import math
import glob
import requests
from datetime import datetime
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
TELEGRAM_CHAT_ID = os.getenv("TELEGRAM_CHAT_ID")

LOG_DIR = "logs"
POLL_INTERVAL = 10
IDLE_TIMEOUT = 300

TOTAL_REGIONS = 54
STATE_FILE = "watcher_state.json"

# ---------------------------------------------------------------------------
# Slug ↔ Bulgarian name mapping (same order as REGION_ORDER)
# Used to key estimates loaded from time_estimates.xlsx
# ---------------------------------------------------------------------------

_SLUG_ORDER = [
    "oblast-blagoevgrad", "oblast-burgas", "oblast-varna", "oblast-veliko-tarnovo",
    "oblast-vidin", "oblast-vratsa", "oblast-gabrovo", "oblast-dobrich",
    "oblast-kardzhali", "oblast-kyustendil", "oblast-lovech", "oblast-montana",
    "oblast-pazardzhik", "oblast-pernik", "oblast-pleven", "oblast-plovdiv",
    "oblast-razgrad", "oblast-ruse", "oblast-silistra", "oblast-sliven",
    "oblast-smolyan", "oblast-sofiya", "oblast-stara-zagora", "oblast-targovishte",
    "oblast-haskovo", "oblast-shumen", "oblast-yambol",
    "grad-blagoevgrad", "grad-burgas", "grad-varna", "grad-veliko-tarnovo",
    "grad-vidin", "grad-vratsa", "grad-gabrovo", "grad-dobrich",
    "grad-kardzhali", "grad-kyustendil", "grad-lovech", "grad-montana",
    "grad-pazardzhik", "grad-pernik", "grad-pleven", "grad-plovdiv",
    "grad-razgrad", "grad-ruse", "grad-silistra", "grad-sliven",
    "grad-smolyan", "grad-sofiya", "grad-stara-zagora", "grad-targovishte",
    "grad-haskovo", "grad-shumen", "grad-yambol",
]

# Same order, Bulgarian names as they appear in the xlsx
_BG_NAME_ORDER = [
    "област Благоевград", "област Бургас", "област Варна", "област Велико Търново",
    "област Видин", "област Враца", "област Габрово", "област Добрич",
    "област Кърджали", "област Кюстендил", "област Ловеч", "област Монтана",
    "област Пазарджик", "област Перник", "област Плевен", "област Пловдив",
    "област Разград", "област Русе", "област Силистра", "област Сливен",
    "област Смолян", "област София", "област Стара Загора", "област Търговище",
    "област Хасково", "област Шумен", "област Ямбол",
    "град Благоевград", "град Бургас", "град Варна", "град Велико Търново",
    "град Видин", "град Враца", "град Габрово", "град Добрич",
    "град Кърджали", "град Кюстендил", "град Ловеч", "град Монтана",
    "град Пазарджик", "град Перник", "град Плевен", "град Пловдив",
    "град Разград", "град Русе", "град Силистра", "град Сливен",
    "град Смолян", "град София", "град Стара Загора", "град Търговище",
    "град Хасково", "град Шумен", "град Ямбол",
]

# region_name → slug  (used to normalise names that come back from the xlsx)
_NAME_TO_SLUG = dict(zip(_BG_NAME_ORDER, _SLUG_ORDER))

# ---------------------------------------------------------------------------
# Load time estimates from xlsx at startup
# Returns: { "prodazhbi": { slug: "Xh Ym" }, "naemi": { slug: "Xh Ym" } }
# ---------------------------------------------------------------------------

XLSX_PATH = "monitoring/time_estimates.xlsx"

# Sheet name in xlsx → transaction type key used in state
_SHEET_TO_TX = {
    "Prodazhbi": "prodazhbi",
    "Naemi":     "naemi",
}


def load_estimates_from_xlsx(path: str) -> dict[str, dict[str, str]]:
    """
    Read time estimates from both sheets.
    Returns { "prodazhbi": {slug: est_str}, "naemi": {slug: est_str} }
    Falls back to empty dicts if the file is missing or unreadable.
    """
    result = {"prodazhbi": {}, "naemi": {}}
    if not os.path.exists(path):
        print(f"[watcher] Warning: {path} not found — estimates unavailable")
        return result
    try:
        wb = load_workbook(path, read_only=True)
        for sheet_name, tx_key in _SHEET_TO_TX.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                bg_name, _, _, est_str = (list(row) + [None, None, None, None])[:4]
                if not bg_name or not est_str:
                    continue
                bg_name = str(bg_name).strip()
                slug = _NAME_TO_SLUG.get(bg_name)
                if slug and est_str is not None:
                    if isinstance(est_str, str):
                        # Plain text value e.g. "1h 25m" — use as-is
                        result[tx_key][slug] = est_str
                    else:
                        # openpyxl returns time-formatted cells as timedelta
                        from datetime import timedelta
                        if isinstance(est_str, timedelta):
                            total_seconds = int(est_str.total_seconds())
                        else:
                            total_seconds = round(float(est_str) * 24 * 3600)
                        h, remainder = divmod(total_seconds, 3600)
                        m, s = divmod(remainder, 60)
                        if h > 0:
                            result[tx_key][slug] = f"{h}h {m:02d}m" if s == 0 else f"{h}h {m:02d}m {s:02d}s"
                        elif m > 0:
                            result[tx_key][slug] = f"{m}m" if s == 0 else f"{m}m {s:02d}s"
                        else:
                            result[tx_key][slug] = f"{s}s"
    except Exception as e:
        print(f"[watcher] Warning: could not load estimates from xlsx: {e}")
    return result


# Loaded once at import time; both sheets, keyed by slug
# Structure: REGION_ESTIMATES["prodazhbi"]["oblast-burgas"] → "6h 41m"
REGION_ESTIMATES = load_estimates_from_xlsx(XLSX_PATH)

# Ordered list of region slugs (same order the scraper visits them)
REGION_ORDER = _SLUG_ORDER


def get_region_estimate(slug: str, transaction_type: str) -> str:
    """
    Return the xlsx estimate string for the given slug + transaction type.
    E.g. "6h 41m" or "<1m".  Returns "" if not found.
    """
    tx = transaction_type or "prodazhbi"
    return REGION_ESTIMATES.get(tx, {}).get(slug, "")


# ---------------------------------------------------------------------------
# Regex patterns — matched against exact log line formats
# ---------------------------------------------------------------------------

# "2026-04-06 09:14:41"
_TIMESTAMP_PATTERN = re.compile(
    r"^(\d{4}-\d{2}-\d{2}\s+\d{2}:\d{2}:\d{2})"
)

# "===== Transaction type: prodazhbi ====="
_TRANSACTION_PATTERN = re.compile(
    r"Transaction type:\s*(\S+)"
)

# "===== Region: Благоевград (oblast-blagoevgrad) | prodazhbi ====="
_REGION_PATTERN = re.compile(
    r"Region:\s*(.+?)\s*\((.+?)\)\s*\|"
)

# "Region 1/76: Благоевград (oblast-blagoevgrad)"
_REGION_NUM_PATTERN = re.compile(
    r"Region\s+(\d+)/(\d+):\s+(.+?)\s*\("
)

# "Pre-flight result for oblast-blagoevgrad/dvustaen: CAPPED"
# "Pre-flight result for oblast-burgas/ednostaen [0-5000000]: CAPPED"
_PREFLIGHT_RESULT_PATTERN = re.compile(
    r"Pre-flight result for (.+?):\s*(CAPPED|under cap)"
)

# "Pre-flight complete. Capped: ['ednostaen', 'dvustaen']"
_PREFLIGHT_COMPLETE_PATTERN = re.compile(
    r"Pre-flight complete\. Capped:\s*(\[.*?\])"
)

# "--- Scraping: oblast-blagoevgrad/ednostaen | prodazhbi | from page 1 ---"
# "--- Scraping: grad-burgas/dvustaen [5000001-10000000] | prodazhbi | from page 1 ---"
_SCRAPING_START_PATTERN = re.compile(
    r"--- Scraping:\s*\S+?/(\S+?)(?:\s*\[.*?\])?\s*\|"
)

# "Saved 38 valid listings from page 3"
_SAVED_LISTINGS_PATTERN = re.compile(
    r"Saved\s+(\d+)\s+valid listings from page\s+(\d+)"
)

# Real errors we care about — these go into the per-region error buffer
# "All 2 attempts failed for URL: ..."
_ALL_ATTEMPTS_FAILED_PATTERN = re.compile(
    r"All \d+ attempts failed for URL:\s*(\S+)"
)

# "Pre-flight fetch failed for oblast-burgas/ednostaen [312501-625000] — assuming no cap."
_PREFLIGHT_FETCH_FAILED_PATTERN = re.compile(
    r"Pre-flight fetch failed for (.+?)\s*—"
)

# "Connection error (attempt 1/2)"
_CONNECTION_ERROR_PATTERN = re.compile(
    r"Connection error"
)

# "Fetch returned None on page N for slug [range]. Stopping." — natural end, suppress
_FETCH_NONE_PATTERN = re.compile(
    r"Fetch returned None on page"
)

# "Cap likely hit for ... — will split further."  — informational, suppress
_CAP_LIKELY_HIT_PATTERN = re.compile(
    r"Cap likely hit for"
)

# "Soft-block suspected" — handled by retry logic, suppress
_SOFT_BLOCK_PATTERN = re.compile(
    r"Soft-block suspected"
)

# "Scraping run started"
_RUN_STARTED_PATTERN = re.compile(
    r"Scraping run started"
)

# "Run finished" or "run finished"
_RUN_FINISHED_PATTERN = re.compile(
    r"[Rr]un finished"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_log_timestamp(line: str) -> float | None:
    """Extract timestamp from a log line and return as epoch seconds."""
    match = _TIMESTAMP_PATTERN.match(line)
    if match:
        try:
            dt = datetime.strptime(match.group(1), "%Y-%m-%d %H:%M:%S")
            return dt.timestamp()
        except ValueError:
            return None
    return None


def extract_time_str(line: str) -> str:
    """Extract HH:MM:SS from a log line for use in notifications."""
    match = _TIMESTAMP_PATTERN.match(line)
    if match:
        return match.group(1).split(" ")[1]
    return datetime.now().strftime("%H:%M:%S")


def fmt_duration(seconds: float) -> str:
    seconds = int(seconds)
    if seconds < 60:
        return f"{seconds}s"
    elif seconds < 3600:
        m, s = divmod(seconds, 60)
        return f"{m}m {s}s"
    else:
        h, remainder = divmod(seconds, 3600)
        m, s = divmod(remainder, 60)
        return f"{h}h {m}m {s}s"


def send_telegram(message: str):
    if not TELEGRAM_TOKEN:
        print(f"[watcher] {message}")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
    try:
        requests.post(url, data={
            "chat_id": TELEGRAM_CHAT_ID,
            "text": message
        }, timeout=10)
    except Exception as e:
        print(f"[watcher] Telegram error: {e}")


def extract_base_property_type(slug_part: str) -> str:
    """Strip price range suffix like ' [0-312500]' from a property type string."""
    return re.sub(r"\s*\[.*?\]", "", slug_part).strip()


def format_error_bullet(raw_line: str) -> str | None:
    """
    Convert a raw log error/warning line into a compact bullet string.
    Returns None if the line should be suppressed entirely.
    """
    if _FETCH_NONE_PATTERN.search(raw_line):
        return None
    if _CAP_LIKELY_HIT_PATTERN.search(raw_line):
        return None
    if _SOFT_BLOCK_PATTERN.search(raw_line):
        return None
    if "Bucket" in raw_line and "splitting" in raw_line:
        return None

    m = _ALL_ATTEMPTS_FAILED_PATTERN.search(raw_line)
    if m:
        url = m.group(1)
        slug_match = re.search(r"/(?:prodazhbi|naemi)/(.+?)(?:\?|/p-|$)", url)
        price_match = re.search(r"price_min=(\d+)&price_max=(\d+)", url)
        slug = slug_match.group(1) if slug_match else url
        if price_match:
            return f"Soft-block — {slug} [{price_match.group(1)}–{price_match.group(2)}]"
        return f"Soft-block — {slug}"

    m = _PREFLIGHT_FETCH_FAILED_PATTERN.search(raw_line)
    if m:
        return f"Pre-flight failed — {m.group(1).strip()}"

    if _CONNECTION_ERROR_PATTERN.search(raw_line):
        return "Connection error"

    return None


# ---------------------------------------------------------------------------
# State persistence
# ---------------------------------------------------------------------------

def load_state() -> dict:
    if os.path.exists(STATE_FILE):
        try:
            with open(STATE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return default_state()


def default_state() -> dict:
    return {
        "region_counter":   0,
        "region_name":      None,
        "region_slug":      None,
        "region_listings":  0,
        "region_start_ts":  None,
        "total_listings":   0,
        "region_times":     [],
        "run_start_ts":     None,
        "log_path":         None,
        "position":         0,
        # current transaction type — updated when scraper announces it
        "transaction_type": "prodazhbi",
        # per-region error buffer — cleared on each new region
        "region_errors":    [],
        # preflight state for the current region
        "preflight_capped": [],
        "preflight_under":  0,
        # track whether we've sent the "starting scrape" message this region
        "scrape_started_sent": False,
        # property type tracking within a region
        "scrape_total_types": 0,        # total unique base types (capped + under)
        "scrape_seen_types":  [],       # base types already seen during scraping
        "scrape_task_counter": 0,       # counter for notifications
    }


def save_state(state: dict):
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False)
    except IOError as e:
        print(f"[watcher] Could not save state: {e}")


def get_latest_log():
    files = glob.glob(os.path.join(LOG_DIR, "*.log"))
    if not files:
        return None
    return max(files, key=os.path.getmtime)


# ---------------------------------------------------------------------------
# Region finalisation — send error summary + completion message
# ---------------------------------------------------------------------------

def finalise_region(state: dict, now_ts: float, time_str: str):
    """Send the error summary (if any) and the ✅ completion message for the current region."""
    region_name = state["region_name"]
    if not region_name:
        return

    duration = now_ts - state["region_start_ts"] if state["region_start_ts"] else 0
    state["region_times"].append(duration)
    state["total_listings"] += state["region_listings"]
    pages = math.ceil(state["region_listings"] / 40) if state["region_listings"] > 0 else 0

    # Send error summary first, if there were real failures
    errors = state.get("region_errors", [])
    if errors:
        seen = set()
        unique_errors = []
        for e in errors:
            if e not in seen:
                seen.add(e)
                unique_errors.append(e)
        bullet_lines = "\n".join(f"• {e}" for e in unique_errors)
        send_telegram(
            f"⚠️ {len(unique_errors)} fetch failure(s) in {region_name}:\n"
            f"{bullet_lines}\n"
            f"🕐 {time_str}"
        )

    send_telegram(
        f"✅ Finished — {region_name}: "
        f"{state['region_listings']} listings / {pages} pages "
        f"in {fmt_duration(duration)}\n"
        f"🕐 {time_str}"
    )

    # Reset per-region state
    state["region_listings"] = 0
    state["region_errors"] = []
    state["preflight_capped"] = []
    state["preflight_under"] = 0
    state["scrape_started_sent"] = False
    state["scrape_total_types"] = 0
    state["scrape_seen_types"] = []
    state["scrape_task_counter"] = 0
    state["region_start_ts"] = now_ts


# ---------------------------------------------------------------------------
# Main watch loop
# ---------------------------------------------------------------------------

def watch():
    print("[watcher] Started")

    state = load_state()

    log_path = state.get("log_path")
    position = state.get("position", 0)
    last_growth = time.time()

    while True:
        latest = get_latest_log()

        if latest != log_path:
            if latest:
                print(f"[watcher] Watching: {latest}")
                log_path = latest
                if state.get("log_path") != latest:
                    position = 0
                last_growth = time.time()
            else:
                time.sleep(POLL_INTERVAL)
                continue

        try:
            with open(log_path, "rb") as f:
                f.seek(position)
                raw = f.read()
                position = f.tell()
        except Exception:
            time.sleep(POLL_INTERVAL)
            continue

        if raw:
            last_growth = time.time()

        # Decode mixed-encoding log
        try:
            text = raw.decode("utf-8")
        except UnicodeDecodeError:
            text = raw.decode("latin-1")

        new_lines = text.split("\n")

        for line in new_lines:
            line = line.strip()
            if not line:
                continue

            line_ts = parse_log_timestamp(line)
            now_ts = line_ts or time.time()
            time_str = extract_time_str(line)
            upper = line.upper()

            # ------------------------------------------------------------------
            # 1. Run started
            # ------------------------------------------------------------------
            if _RUN_STARTED_PATTERN.search(line):
                state["run_start_ts"] = now_ts
                send_telegram(
                    f"🟢 Scraper started\n"
                    f"🕐 {time_str}"
                )

            # ------------------------------------------------------------------
            # 2. Transaction type announcement
            # ------------------------------------------------------------------
            elif "===== TRANSACTION TYPE:" in upper:
                m = _TRANSACTION_PATTERN.search(line)
                tx = m.group(1) if m else "unknown"
                state["region_counter"] = 0
                state["transaction_type"] = tx
                send_telegram(
                    f"📋 Starting: {tx}\n"
                    f"🕐 {time_str}"
                )

            # ------------------------------------------------------------------
            # 2b. "Region 17/76: ..." line — capture counter
            # ------------------------------------------------------------------
            elif _REGION_NUM_PATTERN.search(line) and "===== REGION:" not in upper:
                m_num = _REGION_NUM_PATTERN.search(line)
                state["_pending_region_num"] = int(m_num.group(1))
                state["_pending_region_total"] = int(m_num.group(2))

            # ------------------------------------------------------------------
            # 3. New region starts — finalise previous, announce new
            # ------------------------------------------------------------------
            elif "===== REGION:" in upper:
                # Finalise previous region if there was one
                if state["region_name"] and state["region_start_ts"]:
                    finalise_region(state, now_ts, time_str)

                # Parse region name AND slug
                m_name = _REGION_PATTERN.search(line)

                # Use pending region number if captured from "Region N/N" line
                if state.get("_pending_region_num"):
                    state["region_counter"] = state.pop("_pending_region_num")
                    state.pop("_pending_region_total", None)
                elif m_name:
                    state["region_counter"] += 1
                else:
                    state["region_counter"] += 1

                region_name = m_name.group(1) if m_name else "Unknown"
                region_slug = m_name.group(2) if m_name else None

                state["region_name"] = region_name
                state["region_slug"] = region_slug
                state["region_start_ts"] = now_ts
                state["region_errors"] = []
                state["preflight_capped"] = []
                state["preflight_under"] = 0
                state["scrape_started_sent"] = False
                state["scrape_total_types"] = 0
                state["scrape_seen_types"] = []
                state["scrape_task_counter"] = 0

                pct = round(state["region_counter"] / TOTAL_REGIONS * 100, 1)

                # Estimate from xlsx for this specific region + transaction type
                est_str = ""
                if region_slug:
                    tx = state.get("transaction_type", "prodazhbi")
                    est_val = get_region_estimate(region_slug, tx)
                    if est_val:
                        est_str = f" | ⏳ Est: {est_val}"

                send_telegram(
                    f"📍 Region {state['region_counter']}/{TOTAL_REGIONS} — "
                    f"{region_name} ({pct}%){est_str}\n"
                    f"🕐 {time_str}"
                )

            # ------------------------------------------------------------------
            # 4. Pre-flight individual result — accumulate, don't send yet
            # ------------------------------------------------------------------
            elif _PREFLIGHT_RESULT_PATTERN.search(line):
                m = _PREFLIGHT_RESULT_PATTERN.search(line)
                if m:
                    slug_part = m.group(1).strip()
                    result = m.group(2)
                    # Extract base property type (strip price range)
                    prop_type = slug_part.split("/")[-1] if "/" in slug_part else slug_part
                    prop_type = extract_base_property_type(prop_type)
                    if result == "CAPPED":
                        if prop_type not in state["preflight_capped"]:
                            state["preflight_capped"].append(prop_type)
                    else:
                        # Only count unique base types for "under cap"
                        # (capped types also get preflight results for sub-ranges,
                        #  but we only count the base type once)
                        state["preflight_under"] = state.get("preflight_under", 0) + 1

            # ------------------------------------------------------------------
            # 5. Pre-flight complete — send the summary message
            # ------------------------------------------------------------------
            elif _PREFLIGHT_COMPLETE_PATTERN.search(line):
                capped = state.get("preflight_capped", [])
                under = state.get("preflight_under", 0)
                total_types = len(capped) + under
                state["scrape_total_types"] = total_types

                if capped:
                    capped_str = ", ".join(capped)
                    send_telegram(
                        f"🔍 Pre-flight done — {len(capped)} capped: {capped_str} | "
                        f"{under} under cap ({total_types} types total)\n"
                        f"🕐 {time_str}"
                    )
                else:
                    send_telegram(
                        f"🔍 Pre-flight done — all {under} under cap ({total_types} types total)\n"
                        f"🕐 {time_str}"
                    )

            # ------------------------------------------------------------------
            # 6. Scraping starts — track property type progress
            # ------------------------------------------------------------------
            elif _SCRAPING_START_PATTERN.search(line):
                m_scrape = _SCRAPING_START_PATTERN.search(line)
                base_type = m_scrape.group(1) if m_scrape else None

                # Send "Starting scrape..." once per region
                if not state.get("scrape_started_sent"):
                    state["scrape_started_sent"] = True
                    send_telegram(
                        f"⛏️ Starting scrape...\n"
                        f"🕐 {time_str}"
                    )

                # Track unique base property types and send per-type notification
                if base_type:
                    seen = state.get("scrape_seen_types", [])
                    if base_type not in seen:
                        seen.append(base_type)
                        state["scrape_seen_types"] = seen
                        state["scrape_task_counter"] = len(seen)
                        # Use preflight total as denominator, but never let it be
                        # lower than the number of types we've actually seen so far
                        preflight_total = state.get("scrape_total_types", 0)
                        total = max(preflight_total, len(seen))
                        send_telegram(
                            f"⛏️ Scraping type {state['scrape_task_counter']}/{total} — {base_type}\n"
                            f"🕐 {time_str}"
                        )

            # ------------------------------------------------------------------
            # 7. Listings saved — accumulate count silently
            # ------------------------------------------------------------------
            elif _SAVED_LISTINGS_PATTERN.search(line):
                m = _SAVED_LISTINGS_PATTERN.search(line)
                if m:
                    state["region_listings"] = state.get("region_listings", 0) + int(m.group(1))

            # ------------------------------------------------------------------
            # 8. Run finished
            # ------------------------------------------------------------------
            elif _RUN_FINISHED_PATTERN.search(line):
                # Finalise last region
                if state["region_name"] and state["region_start_ts"]:
                    finalise_region(state, now_ts, time_str)

                total = state["total_listings"]
                total_pages = math.ceil(total / 40) if total > 0 else 0
                run_dur = fmt_duration(now_ts - state["run_start_ts"]) if state["run_start_ts"] else "unknown"
                regions_done = len(state["region_times"])

                send_telegram(
                    f"🏁 Run complete!\n"
                    f"📊 {total} listings / {total_pages} pages\n"
                    f"🗺 {regions_done} regions\n"
                    f"⏱ {run_dur}\n"
                    f"🕐 {time_str}"
                )

                if os.path.exists(STATE_FILE):
                    os.remove(STATE_FILE)
                return

            # ------------------------------------------------------------------
            # 9. Errors — collect into per-region buffer (sent at region end)
            # ------------------------------------------------------------------
            elif "[ERROR]" in upper or "[WARNING]" in upper:
                bullet = format_error_bullet(line)
                if bullet:
                    state.setdefault("region_errors", []).append(bullet)

        # Save state after each batch
        save_state({
            **state,
            "log_path": log_path,
            "position": position,
        })

        # Idle watchdog
        if time.time() - last_growth > IDLE_TIMEOUT:
            send_telegram(
                f"⚠️ No log activity for {IDLE_TIMEOUT // 60} min — possible crash\n"
                f"🕐 {datetime.now().strftime('%H:%M:%S')}"
            )
            last_growth = time.time()

        time.sleep(POLL_INTERVAL)


if __name__ == "__main__":
    watch()
